import csv
import openpyxl
import os
from .models import Users, studentProfile, question, quesFile, studentMark

root_path = os.path.dirname(os.path.abspath(__file__))

def file_to_db(filename, client_name):
    '''main function: write file to database'''
    print(filename[-3:] + 'sdaasd')
    if filename[-3:] == 'csv':
        data = read_csv(filename)
        write_db(data, client_name)        
    elif filename[-4:] == 'xlsx' or filename[-3:] == 'xls':
        read_xl(filename)
        write_db(data, client_name)
    else:
        print('error')
    
def read_csv(filename):
    '''read csv file'''
    with open(root_path + '/static/onlinetest/docs/' + filename) as csvfile:
        spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
        data = {}
        data_row = []
        i = 0
        for row in spamreader:
            data[i] = {}
            j = 0
            data_row = row[0].split(',')
            for row_cell in data_row:
                data[i][j] = row_cell
                j += 1
            i += 1
    data['filename'] = filename
        #print(data)
    return data

    
def read_xl(filename):
    '''read excel file'''
    wb = openpyxl.load_workbook(root_path + '/static/onlinetest/docs/' + filename)
    print(wb.get_sheet_names())
    anotherSheet = wb.active
    ws = wb.worksheets[0]
    #print(ws.read())
    #total sheets
    data = {}
    for i in wb.worksheets:
        #all rows and columns
        row_count = i.max_row
        col_count = i.max_column
        print(row_count, col_count)
        for row in range(2,row_count):
            data[row-2] = {}
            for col in range(2,col_count):
                data[row-2][col-2] = i.cell(column=col, row=row).value
    data['filename'] = filename
    #print(data)
    return data

def write_db(data, client_name):
    '''write to database'''
    #ques_paper=quesFile.objects.get(ques_paper_id=filename, client=client_name)
    for i in data:
        if str(i) == 'filename':
            break
        ques=question.objects.create(
            question_id=data[i].get(0),
            question=data[i].get(1),
            option1=data[i].get(2),
            option2=data[i].get(3),
            option3=data[i].get(4),
            option4=data[i].get(5),
            answer=data[i].get(6),
            questionType=data[i].get(7),
    )
    '''for i in data:
        print(data[i].get(0))
        print(data[i].get(1))
        print(data[i].get(2))
        print(data[i].get(3))
        print(data[i].get(4))
        print(data[i].get(5))
        print(data[i].get(6))
        print(data[i].get(7))
     '''   
        
#read_xl('quesformat.xlsx')

#file_to_db(read_csv('file1.csv','yoyo')

